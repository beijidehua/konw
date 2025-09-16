# dvadmin/utils/export.py

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from datetime import datetime


class ExportMixin:
    """导出功能混入类"""

    def export_excel(self, request, *args, **kwargs):
        """导出Excel"""
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.model._meta.verbose_name

        # 表头样式
        header_font = Font(name='微软雅黑', bold=True)
        header_fill = PatternFill('solid', fgColor='1874CD')
        align = Alignment(horizontal='center', vertical='center')

        # 获取要导出的数据
        queryset = self.filter_queryset(self.get_queryset())
        header = self.get_export_headers()

        # 写入表头
        for col, name in enumerate(header.values(), start=1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align

        # 写入数据
        for row, instance in enumerate(queryset, start=2):
            for col, field in enumerate(header.keys(), start=1):
                value = self.get_export_field_value(instance, field)
                ws.cell(row=row, column=col, value=value)

        # 生成响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{self.model._meta.verbose_name}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def get_export_headers(self):
        """获取导出表头"""
        raise NotImplementedError('需要实现 get_export_headers 方法')

    def get_export_field_value(self, instance, field):
        """获取导出字段值"""
        raise NotImplementedError('需要实现 get_export_field_value 方法')
